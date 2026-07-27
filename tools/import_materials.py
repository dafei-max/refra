from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
IMAGE_DIR = ROOT / "素材资产库图片素材"


ALIASES = {
    "number": ["素材编号", "编号", "number", "Number"],
    "type": ["类型", "素材类型", "type", "Type"],
    "image": ["图片image", "图片", "image", "Image", "素材图片", "图片URL"],
    "category": ["适用品类", "品类", "category", "Category"],
    "reference_description": ["参考描述", "Reference", "reference", "可参考", "描述"],
}


def pick(row: dict, key: str) -> str:
    for name in ALIASES[key]:
        value = row.get(name)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def normalize_number(value: str) -> str:
    return value.strip().replace("-", "_")


def normalize_image(value: str, number: str) -> str:
    if value.startswith(("http://", "https://", "/assets/", "/uploads/")):
        return value
    if value:
        candidate = IMAGE_DIR / value
        if candidate.exists():
            return f"/assets/{candidate.name}"
    if number and (IMAGE_DIR / f"{number}.png").exists():
        return f"/assets/{number}.png"
    return value


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: import_materials.py <xlsx>")
    workbook = Path(sys.argv[1])
    wb = openpyxl.load_workbook(workbook, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    materials = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        source = dict(zip(headers, row))
        number = normalize_number(pick(source, "number") or f"IMPORT_{idx}")
        material = {
            "number": number,
            "type": pick(source, "type"),
            "image": normalize_image(pick(source, "image"), number),
            "category": pick(source, "category"),
            "reference_description": pick(source, "reference_description"),
        }
        if material["number"] and material["type"]:
            materials.append(material)
    print(json.dumps(materials, ensure_ascii=False))


if __name__ == "__main__":
    main()
