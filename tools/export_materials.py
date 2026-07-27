from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "素材资产库.xlsx"
IMAGE_DIR = ROOT / "素材资产库图片素材"
OUTPUT = ROOT / "data" / "materials.json"


def normalize_number(value: str | None) -> str:
    if not value:
        return ""
    return str(value).strip().replace("-", "_")


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    materials = []

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        source = dict(zip(headers, row))
        number = normalize_number(source.get("素材编号"))
        local_file = IMAGE_DIR / f"{number}.png"
        local_image = f"/assets/{number}.png" if local_file.exists() else None

        materials.append(
            {
                "number": number,
                "type": source.get("素材类型") or "",
                "image": local_image or source.get("图片URL") or "",
                "category": source.get("适用品类") or "",
                "reference_description": source.get("参考描述")
                or source.get("可参考")
                or source.get("优点")
                or source.get("风格")
                or "",
            }
        )

    payload = {
        "source": str(WORKBOOK.name),
        "image_dir": str(IMAGE_DIR.name),
        "count": len(materials),
        "materials": materials,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT} with {len(materials)} materials")


if __name__ == "__main__":
    main()
